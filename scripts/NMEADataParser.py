import datetime
import getopt
import sys
from os.path import exists
from pathlib import Path
import openpyxl
from openpyxl.styles import Font
from pynmeagps import NMEAReader

KNOTS_TO_MS = 0.514444
MPH_TO_MS = 0.44704
KPH_TO_MS = 0.277777778
MS_TO_MS = 1

units = {
    "kn": KNOTS_TO_MS,
    "kph": MPH_TO_MS,
    "mph": KPH_TO_MS,
    "ms": MS_TO_MS,
}

scriptname = 'NMEADataParser.py'


# Convert speed to meters per second
def convertSpeed(speed, unit):
    return speed * units[unit]


# Parse data from inputfile and save them to XLSX outputfile
def parseData(inputfile, outputfile, unit):
    stream = open(inputfile, 'rb')
    nmr = NMEAReader(stream, nmeaonly=False)

    # Create a new XLSX workbook
    workbook = openpyxl.Workbook()

    # Select the active worksheet
    worksheet = workbook.active

    # Set column headers and bold font
    headers = ['latitude', 'longitude', 'speed', 'time']
    for col_num, header in enumerate(headers, 1):
        cell = worksheet.cell(row=1, column=col_num)
        cell.value = header
        cell.font = Font(bold=True)

    # Write data to the worksheet with increasing row index
    row_num = 2
    for (raw_data, parsed_data) in nmr:
        msg = NMEAReader.parse(raw_data.decode())
        if msg.msgID == 'RMC':
            worksheet.cell(row=row_num, column=1).value = float(str(msg.lat))
            worksheet.cell(row=row_num, column=2).value = float(str(msg.lon))
            worksheet.cell(row=row_num, column=3).value = convertSpeed(msg.spd, unit)

            # Convert UTC time to Posix time
            date = (str(msg.date) + " " + (str(msg.time)))
            date_time = datetime.datetime.strptime(date, '%Y-%m-%d %H:%M:%S')
            timestamp = date_time.timestamp()

            worksheet.cell(row=row_num, column=4).value = timestamp
            row_num += 1

    # Save the workbook to a file
    workbook.save(outputfile)

    # Close the input stream
    stream.close()


# Print help dialog
def printOptions():
    print('usage:', scriptname, '-i <inputfile.nmea> -o <outputfile.xlsx> -u <unit>')
    print('\t-h help')
    print('\t-i <inputfile> in .nmea / .txt format')
    print('\t-o <outputfile> in .xlsx format')
    print('\t-u <unit> ms for meters per second, kn for knots, kph for kilometers per hour, mph for miles per hour')


# Print missing parameter warning
def printMissing():
    print('Missing arguments')
    print('Wrong input, use -h for more info')


def main(argv):
    inputfile = ''
    outputfile = ''
    unit = ''

    try:
        opts, args = getopt.getopt(argv, "hi:o:u:", ["ifile=", "ofile=", "unit="])
    except getopt.GetoptError:
        printMissing()
        sys.exit(1)

    # Load input parameters
    for opt, arg in opts:
        if opt == '-h':
            printOptions()
            sys.exit(0)

        elif opt in ("-i", "--ifile"):
            inputfile = arg
            fileFormat = Path(inputfile).suffix
            if fileFormat != '.nmea' and fileFormat != '.txt':
                print("Error in [<inputfile>] - File format", fileFormat,
                      "is not supported\nFor more info use -h parameter")
                sys.exit(1)
            if not exists(inputfile):
                print("Error in [<inputfile>] - Input file", inputfile, "doesn't exist\nFor more info use -h parameter")
                sys.exit(1)

        elif opt in ("-o", "--ofile"):
            outputfile = arg
            fileFormat = Path(outputfile).suffix
            if fileFormat != '.xlsx':
                print("Error in [<outputfile>] - File format", fileFormat,
                      "is not supported\nFor more info use -h parameter")
                sys.exit(1)

        elif opt in ("-u", "--unit"):
            unit = arg
            if unit not in units.keys():
                print("Error in [<unit>] - Unit", unit, "is not supported\nFor more info use -h parameter")
                sys.exit(1)

    # Number of parameters is invalid
    if len(argv) != 6:
        printMissing()
        sys.exit(1)

    else:
        print('Parsing...')
        parseData(inputfile, outputfile, unit)

    sys.exit("File successfully parsed")


if __name__ == "__main__":
    main(sys.argv[1:])
